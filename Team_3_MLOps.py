
#Import libraries
from multiprocessing import parent_process
import streamlit as st
import pandas as pd
import numpy as np
from PIL import Image
import joblib
from io import BytesIO
import base64
import openpyxl
# from to_excel_v1 import to_excel_auto_width

from openpyxl.utils import get_column_letter

from all_lists import *

#load the model from disk
filename = "catb_model_24outbaski.sav"
model = joblib.load(filename)

#Setting Application title
st.title('Developer Salary Prediction Model App Team-3')

#Setting Application description
st.markdown("""
This Streamlit app is made to predict Developers Salaries based on Stack Overflow 2018 Developers Survey.
The application is functional for both online prediction and batch data prediction. \n
""")
st.markdown("<h3></h3>", unsafe_allow_html=True)

#Setting Application sidebar default
image = Image.open("App.png")
add_selectbox = st.sidebar.selectbox(
"How would you like to predict?", ("Online", "Batch"))
st.sidebar.info('This app is created to predict Developers Salaries based on Stack Overflow 2018 Developers Survey')
st.sidebar.image(image)

if add_selectbox == "Online":
    st.info("Please Input data below")

    #Based on our optimal features selection
    st.subheader("Developer's Informations")

    country = st.selectbox(label = "Choose a country", options = Country_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    yearsCodingProf = st.selectbox(label = "For how many years have you coded professionally (as a part of your work)?", options = YearsCodingProf_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    raceEthnicity = st.selectbox(label = "Which of the following do you identify as?", options = RaceEthnicity_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    age = st.selectbox(label = "Age", options = Age_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    employment = st.selectbox(label = "Which of the following best describes your current employment status?", options = Employment_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    formalEducation = st.selectbox(label = "Which of the following best describes the highest level of formal education that you have completed?", options = FormalEducation_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    assessBenefits6 = st.number_input(label = "Please rank the 'Retirement or pension savings matching' aspects of a job's benefits package, where 1 is most important and 11 is least important.", min_value=1, max_value=11, value=1)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    communicationTools = st.selectbox(label = "Which of the following tools do you use to communicate, coordinate, or share knowledge with your coworkers?", options = CommunicationTools_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    numberMonitors = st.selectbox(label = "How many monitors are set up at your workstation?", options = NumberMonitors_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    devType = st.selectbox(label = "Which of the following describe you?", options = DevType_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    assessBenefits3 = st.number_input(label = "Please rank the 'Health insurance' aspects of a job's benefits package, where 1 is most important and 11 is least important.", min_value=1, max_value=11, value=1)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    lastNewJob = st.selectbox(label = "When was the last time that you took a job with a new employer?", options = LastNewJob_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    student = st.selectbox(label = "Are you currently enrolled in a formal, degree-granting college or university program?", options = Student_list)
    st.markdown("<h3></h3>", unsafe_allow_html=True)
    ide = st.selectbox(label = "Which development environment(s) do you use regularly?", options= ide_list)

    data = {
            'Country_LE': country,
            'YearsCodingProf' : yearsCodingProf,
            'RaceEthnicity_LE' : raceEthnicity,
            'Age' : age,
            'Employment_LE' : employment,
            'FormalEducation' : formalEducation,
            'CommunicationTools_LE' : communicationTools,
            'AssessBenefits6' : assessBenefits6,
            'NumberMonitors' : numberMonitors,
            'DevType_LE' : devType,
            'AssessBenefits3' : assessBenefits3,
            'LastNewJob' : lastNewJob,
            'Student_LE' : student,
            'IDE_LE' : ide  
        
            }

    features_df = pd.DataFrame.from_dict([data])

    st.markdown("<h3></h3>", unsafe_allow_html=True)
    st.write('Overview of input is shown below')
    st.markdown("<h3></h3>", unsafe_allow_html=True)

    st.write(features_df)

    for i in column_list_to_map:
        if i in binary_list: 
            features_df[i][0] = dictionary.get(features_df[i][0])
            # features_df.loc[i, 0] = dictionary.get(features_df.loc[i, 0])
            
        else:
            pass

    st.markdown("<h3></h3>", unsafe_allow_html=True)

    for j in column_list_to_map:
        features_df[j] = pd.to_numeric(features_df[j], downcast='float')

    if st.button('Predict'):
        prediction = model.predict(features_df)
        prediction_df = pd.DataFrame(prediction, columns=["Prediction"])
        st.markdown("<h3></h3>", unsafe_allow_html=True)

        # Round the prediction value and display it in string format
        prediction_value = round(prediction_df["Prediction"][0])
        prediction_value = f"{prediction_value:,}".replace(",", ".")
        info_message = f"Predicted value: $ {prediction_value}"
        st.info(info_message)

else:
    st.subheader("Dataset upload")
    uploaded_file = st.file_uploader("Choose a file")
    if uploaded_file is not None:
        df_batch = pd.read_csv(uploaded_file,encoding= 'utf-8', low_memory=False, index_col=False)
        st.write(df_batch)
        st.markdown("<h3></h3>", unsafe_allow_html=True)
        
        #Preprocess inputs        
        df_new_batch = pd.DataFrame(columns=column_list_to_map)

        for i in column_list_to_map:
            if i in binary_list:
                df_new_batch[i] = df_batch[i].map(dictionary)
            
            else:
                counter = 0
                while counter < df_batch.shape[0]:                    
                    df_new_batch[i][counter] = df_batch[i][counter]
                    counter += 1

        for j in column_list_to_map:
            df_new_batch[j] = pd.to_numeric(df_new_batch[j], downcast='float')

        if st.button('Predict'):
            prediction = model.predict(df_new_batch)

            prediction_df = pd.DataFrame(prediction, columns=["Prediction_$"])

            # Dropping the decimal parts of the data in the "Prediction_$" column and using a dot as the thousands separator
            prediction_df['Prediction_$'] = prediction_df['Prediction_$'].astype(float).map(lambda x: "{:,.0f}".format(x).replace(',', '.'))

            result_df = pd.concat([df_batch, prediction_df], axis=1, ignore_index=False, sort=False)

            st.markdown("<h3></h3>", unsafe_allow_html=True)
            st.subheader('Prediction')
            st.write(result_df)

            # Convert DataFrame to CSV string
            csv = result_df.to_csv(index=False)
            
            # Present the CSV string to the user with the download button
            st.download_button(
                label="Download predictin as a CSV file",
                data=csv,
                file_name='prediction_results.csv',
                mime='text/csv',
                )
            def to_excel_auto_width(df):
                output = BytesIO()
                # Convert DataFrame to an Excel file
                with pd.ExcelWriter(output, engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    
                    # Adjust column widths
                    for col_num, column in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), start=1):
                        max_length = 0
                        for cell in column:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                        
                        # Set the column width in Excel, adding a bit extra space by using +2
                        worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2
            
                output.seek(0)  # Reset the file pointer to the beginning
                return output


            excel_file = to_excel_auto_width(result_df)
            st.download_button(
                label="Download prediction as an Excel file",
                data=excel_file,
                file_name="prediction_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
